from openpyxl import Workbook, load_workbook
import os

class Excel:
   ARQ = "D:\\faculdade\\python\\LTP Python\\arquivos .XLSX\\nomes.xlsx"
   wb = None
   ws = None
   linha = 0

   def __init__(self):
       if os.path.isfile(self.ARQ):
           self.wb = load_workbook(self.ARQ)
           self.ws = self.wb['Planilha com nomes']
           self.linha = self.ws.max_row + 1
       else:
           self.wb = Workbook()
           self.ws = self.wb.active
           self.ws.title = "Planilha com nomes"
           self.ws['A1'] = "Número"
           self.ws['B1'] = "Nome Fornecido"
           self.ws.column_dimensions['B'].width = 40
           self.linha = 2

   def entrada(self):
       # Manipulando Excel
       nome = ""
       A = 1
       B = 2
       while nome != 'FIM':
          print ("Informe um nome [FIM - Termina o programa]: ")
          nome = input()
          if nome == 'FIM':
              continue
          num = str(self.linha-1)
          self.ws.cell(row=self.linha, column=A, value=num)
          self.ws.cell(row=self.linha, column=B, value=nome)
          self.linha += 1

   def saida(self):
       for linha in range(2, self.linha):
           A = 1
           B = 2
           numero = self.ws.cell(column=A, row=linha).value
           nome = self.ws.cell(column=B, row=linha).value
           print("Nome {}: {}".rjust(10).format(numero, nome))
       pass

   def salvar(self):
       self.wb.save(self.ARQ)
       print("Planilha salva em D:/faculdade/python/LTP Python/arquivos .XLSX")


def main():
   flag = True
   excel = Excel()
   while flag:
       print("-" * 30)
       print("1 - Entrada de nomes")
       print("2 - Saída de nomes")
       print("3 - Terminar")
       print("-" * 30)
       opc = int(input("Qual a opção? "))
       if opc == 1:
           excel.entrada()
       elif opc == 2:
           excel.saida()
       elif opc == 3:
           excel.salvar()
           print("Fim de Programa")
           flag = False


if __name__ == "__main__":
   main()
