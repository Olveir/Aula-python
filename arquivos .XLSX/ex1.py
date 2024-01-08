from openpyxl import load_workbook

wb = load_workbook("D:\\faculdade\\python\\LTP Python\\arquivos .XLSX\\alunos.xlsx")

ws = wb['Planilha1']

A = 1
B = 2
C = 3
D = 4

for linha in range(2, 7):
   aluno = ws.cell(column=A, row=linha).value
   nota1 = ws.cell(column=B, row=linha).value
   nota2 = ws.cell(column=C, row=linha).value
   media = ws.cell(column=D, row=linha).value
   print("Aluno:".rjust(10), aluno)
   print("Prova-1:".rjust(10), nota1)
   print("Prova-2:".rjust(10), nota2)
   print("Média:".rjust(8), media)
   if media == 0:
       print("Menção: SR".rjust(12))
   elif media > 0 and media < 2:
       print("Menção: II".rjust(12))
   elif media >= 2 and media < 5:
       print("Menção: MI".rjust(12))
   elif media >= 5 and media < 7:
       print("Menção: MM".rjust(12))
   elif media >= 7 and media < 9:
       print("Menção: MS".rjust(12))
   else:
       print("Menção: SS".rjust(12))
   print("-------------------------------------")
